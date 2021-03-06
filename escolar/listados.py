# -*- encoding: utf-8 -*-
import os
from django.conf import settings
from django.http import HttpResponse
from django.template import Context
from django.template.loader import get_template
from escolar.models import Docente, Curso, Alumno, MatriculaAlumnado, Campo, MatriculaDocentes, SITUACION_DOCENTE, TIPO_MATRICULA_DOCENTE, ItemCampo, DescripcionCampo
from xhtml2pdf import pisa  #INSTALAR ESTA LIBRERIA
from django.templatetags.static import static
from escolar.default_data.images_base64 import LOGO_PROVINCIAL, LOGO_AMPARO
##PARA MANDAR LOGS
# import the logging library
import logging
# Get an instance of a logger
logger = logging.getLogger('django')

##BASE DE DATOS
from django.db.models import Q

from functions import numero_to_letras, convierte_cifra
import csv

##EXCEL

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

FILE_LIST = settings.BASE_DIR+'/test.pdf'
# Convert HTML URIs to absolute system paths so xhtml2pdf can access those resources
def link_callback(uri, rel):
    # use short variable names
    sUrl = settings.STATIC_URL      # Typically /static/
    sRoot = settings.STATIC_ROOT    # Typically /home/userX/project_static/
    mUrl = settings.MEDIA_URL       # Typically /static/media/
    mRoot = settings.MEDIA_ROOT     # Typically /home/userX/project_static/media/BASE_DIR
    
    # convert URIs to absolute system paths
    if uri.startswith(mUrl):
        path = os.path.join(mRoot, uri.replace(mUrl, ""))
    elif uri.startswith(sUrl):
        path = os.path.join(sRoot, uri.replace(sUrl, ""))

    # make sure that file exists
    if not os.path.isfile(path):
            raise Exception(
                    'media URI must start with %s or %s' % \
                    (sUrl, mUrl))
    return path

##listados

def generar_listado_gdipe(request, anio):
    response = HttpResponse(content_type='application/pdf')

    logger.info("Levanto http")    
    response['Content-Disposition'] = 'attachment; filename="listado_gdipe.pdf"'
    cursos = Curso.objects.filter(ciclo=anio)
    lista_matriculas = MatriculaAlumnado.objects.filter(Q(curso__in = cursos))
    logger.info("Recibo cursos" + str(lista_matriculas))
    #Image
    #url = settings.BASE_DIR+ '/static/images/logo_amparo.png'
    #with open(url, "rb") as image_file:
    #    encoded_string = base64.b64encode(image_file.read())
    # Prepare context
    data = {
            'anio':anio,
           'cursos':cursos,
            'matriculas':lista_matriculas,
           }    

    # Render html content through html template with context
    template = get_template('listados/listado_gdipe.html')
    logger.info("Recibo template")
    html  = template.render(Context(data))
    logger.info("Render template")
    # Write PDF to file
    file = open(FILE_LIST, "w+b")
    pisaStatus = pisa.CreatePDF(html, dest=file,
            link_callback = link_callback)

    logger.info("Convierto pdf")
    # Return PDF document through a Django HTTP response
    file.seek(0)
    pdf = file.read()
    file.close()
    logger.info("Meto pdf")
    response.write(pdf)
    # Don't forget to close the file handle
    #BORRO EL ARCHIVO
    if os.path.exists(FILE_LIST):
        try:
            os.remove(FILE_LIST)
        except OSError, e:
            pass
    
    return response

def generar_listado_gdipe_no(request, anio):
    
    cursos = Curso.objects.filter(ciclo=anio)
    
    #Image
    url = static('images/logo_amparo.png')
    with open(url, "rb") as image_file:
        encoded_string = base64.b64encode(image_file.read())
    # Prepare context
    data = {
            'anio':anio,
           'cursos':cursos,
           'logo_64':encoded_string,
           }       

    # Render html content through html template with context
    template = get_template('listados/listado_gdipe.html')
    html  = template.render(Context(data))

    return HttpResponse(html)

def generar_listado_alumnos_pdf(request, id_curso):
    response = HttpResponse(content_type='application/pdf')
    
    try:
        curso = Curso.objects.get(pk=id_curso)  
    except Curso.DoesNotExist:
        raise Http404("El curso no existe")
        
    response['Content-Disposition'] = 'attachment; filename="listado_%s.pdf"' % (curso.curso_abreviado())
    matriculados = MatriculaAlumnado.objects.filter(curso=curso).exclude(activo=False)
    
    # Prepare context
    data = {'curso':curso,
           'matriculados':matriculados,
           }    

    # Render html content through html template with context
    template = get_template('listados/listado_alumnos.html')
    html  = template.render(Context(data))

    # Write PDF to file
    file = open(FILE_LIST, "w+b")
    pisaStatus = pisa.CreatePDF(html, dest=file,
            link_callback = link_callback)

    # Return PDF document through a Django HTTP response
    file.seek(0)
    pdf = file.read()
    file.close()
    
    response.write(pdf)
    # Don't forget to close the file handle
    #BORRO EL ARCHIVO
    if os.path.exists(FILE_LIST):
        try:
            os.remove(FILE_LIST)
        except OSError, e:
            pass
    
    return response


def generar_listado_telefonos_alumnos_pdf(request, id_curso):
    response = HttpResponse(content_type='application/pdf')

    try:
        curso = Curso.objects.get(pk=id_curso)
    except Curso.DoesNotExist:
        raise Http404("El curso no existe")

    response['Content-Disposition'] = 'attachment; filename="listado_tel_%s.pdf"' % (curso.curso_abreviado())
    matriculados = MatriculaAlumnado.objects.filter(curso=curso).exclude(activo=False)

    # Prepare context
    data = {'curso':curso,
           'matriculados':matriculados,
           }

    # Render html content through html template with context
    template = get_template('listados/listado_telefonos_alumnos.html')
    html  = template.render(Context(data))

    # Write PDF to file
    file = open(FILE_LIST, "w+b")
    pisaStatus = pisa.CreatePDF(html, dest=file,
            link_callback = link_callback)

    # Return PDF document through a Django HTTP response
    file.seek(0)
    pdf = file.read()
    file.close()

    response.write(pdf)
    # Don't forget to close the file handle
    #BORRO EL ARCHIVO
    if os.path.exists(FILE_LIST):
        try:
            os.remove(FILE_LIST)
        except OSError, e:
            pass

    return response

def generar_informe_matricula(request, matricula_id, etapa):
    response = HttpResponse(content_type='application/pdf')
    
    try:
        matricula = MatriculaAlumnado.objects.get(pk=matricula_id)
    except MatriculaAlumnado.DoesNotExist:
        raise Http404("El curso no existe")
        
    response['Content-Disposition'] = 'attachment; filename="informe%s.pdf"' % (matricula.alumno)
    matriculados = MatriculaAlumnado.objects.filter(curso=curso).exclude(activo=False)
    
    # Prepare context
    data = {'curso':curso,
           'matriculados':matriculados,
           }    

    # Render html content through html template with context
    template = get_template('informe/_informe.html')
    html  = template.render(Context(data))

    # Write PDF to file
    file = open(FILE_LIST, "w+b")
    pisaStatus = pisa.CreatePDF(html, dest=file,
            link_callback = link_callback)

    # Return PDF document through a Django HTTP response
    file.seek(0)
    pdf = file.read()
    file.close()
    
    response.write(pdf)
    # Don't forget to close the file handle
    #BORRO EL ARCHIVO
    if os.path.exists(FILE_LIST):
        try:
            os.remove(FILE_LIST)
        except OSError, e:
            pass
    
    return response



def informe_matricula(request, id_matricula_alumno, etapa):
    response = HttpResponse(content_type='application/pdf')
            
    response['Content-Disposition'] = 'attachment; filename="informe_test.pdf"'    
    
    # Prepare context
    
    matricula = MatriculaAlumnado.objects.get(pk=id_matricula_alumno)
    descrCampo = DescripcionCampo.objects.filter(matricula_alumno=matricula, semestre=etapa, campo__especial=False)
    descrCampoInstitucionales = DescripcionCampo.objects.filter(matricula_alumno=matricula, semestre=etapa, campo__especial=True)
    
    
    data = {'etapa':int(etapa),
            'matricula':matricula,
            'descrCampo':descrCampo,
            'descrCampoInstitucionales':descrCampoInstitucionales,
            'logo_provincial':LOGO_PROVINCIAL
           }    

    # Render html content through html template with context
    template = get_template('informe/informe.html')
    html  = template.render(Context(data))

    # Write PDF to file
    file = open(FILE_LIST, "w+b")
    pisaStatus = pisa.CreatePDF(html, dest=file,
            link_callback = link_callback)

    # Return PDF document through a Django HTTP response
    file.seek(0)
    pdf = file.read()
    file.close()
    
    response.write(pdf)
    # Don't forget to close the file handle
    #BORRO EL ARCHIVO
    if os.path.exists(FILE_LIST):
        try:
            os.remove(FILE_LIST)
        except OSError, e:
            pass
    
    return response





def generar_base_para_certificados(request, id_curso):
    import locale
    locale.setlocale(locale.LC_ALL,'es_AR.utf8')
    try:
        curso = Curso.objects.get(pk=id_curso)
    except Curso.DoesNotExist:
        raise Http404("El curso no existe")

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="{}.csv"'.format(curso.curso_abreviado())

    matriculados = MatriculaAlumnado.objects.filter(curso=curso).exclude(activo=False)



    writer = csv.writer(response)
    writer.writerow(['Orden','Nombre','DNI','ciudad', 'dpto','provincia','pais','dia','mes','anio'])
    for matricula in matriculados:
        mat_order = "{}".format(matricula.numMatricula).zfill(3)
        alumno = matricula.alumno
        nombre = "{} {}".format(alumno.nombres.upper(), alumno.apellidos.upper())
        dni = "{:,}".format(int(alumno.dni)).replace(',','.')

        try:
            ciudad, dpto = alumno.lugarDeNacimiento.split(' ')
        except:
            ciudad, dpto = "Córdoba", "Capital"

        if (len(dpto)<3):
            dpto = "Capital"
        ciudad=ciudad.title()
        dpto=dpto.title()
        provincia = alumno.provincia.title()
        pais = alumno.pais.title()
        dia = numero_to_letras(int(alumno.fechaDeNacimiento.strftime("%d"))).lower()
        mes = alumno.fechaDeNacimiento.strftime("%B").lower()
        anio = numero_to_letras(int(alumno.fechaDeNacimiento.strftime("%Y"))).lower()

        writer.writerow([mat_order,nombre,dni,ciudad, dpto,provincia,pais,dia,mes,anio])

    return response



def generar_listado_promovidos_pdf(request, anio):

    response = HttpResponse(content_type='application/pdf')
    logger.info("Levanto http")
    response['Content-Disposition'] = 'attachment; filename="listado_promovidos_{}.pdf"'.format(anio)
    cursos = Curso.objects.filter(ciclo=anio, anio=5)
    # Prepare context
    data = {
           'cursos':cursos,
           }

    # Render html content through html template with context
    template = get_template('listados/listado_promovidos.html')

    logger.info("Recibo template")
    html  = template.render(Context(data))
    logger.info("Render template")
    # Write PDF to file
    file = open(FILE_LIST, "w+b")
    pisaStatus = pisa.CreatePDF(html, dest=file,
            link_callback = link_callback)

    logger.info("Convierto pdf")
    # Return PDF document through a Django HTTP response
    file.seek(0)
    pdf = file.read()
    file.close()
    logger.info("Meto pdf")
    response.write(pdf)
    # Don't forget to close the file handle
    #BORRO EL ARCHIVO
    if os.path.exists(FILE_LIST):
        try:
            os.remove(FILE_LIST)
        except OSError, e:
            pass

    return response

def generar_listado_promovidos_xls(request, anio):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;')
    response['Content-Disposition'] = 'attachment; filename=lista_promovidos_{}.xlsx'.format(anio)


    title = Font(name='Arial', size=14, vertAlign="baseline", b=True, )
    subtitle14 = Font(name='Arial', size=14)
    subtitle12 = Font(name='Arial', size=12, b=True)
    thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

    cursos = Curso.objects.filter(ciclo=anio, anio=5)
    # Prepare context
    data = {
           'cursos':cursos,
           }
    workbook = Workbook()
    workbook.remove_sheet(workbook.active)
    for curso in cursos:
        ws = workbook.create_sheet()
        ws.title = "{}-{}".format(anio, curso.sala)
        title_range = ws.merge_cells('A2:C2')
        subtitle_range = ws.merge_cells('A3:C3')
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 13
        ws['A2'].font = title
        ws['A3'].font = subtitle14
        ws['A2'] = "ALUMNOS PROMOVIDOS A PRIMER GRADO DE LA EDUCACIÓN PRIMARIA"
        ws.row_dimensions[2].height = 40
        ws['A2'].alignment = Alignment(horizontal='center',vertical='center')
        ws['A3'] = "Ley de Educación Nacional Nº 26.206 - Ley Provincial Nº 9870"
        ws['A3'].alignment = Alignment(horizontal='center',vertical='center')
        ws['A5'].font = subtitle12
        ws['A5'] = "Sección: '{}'".format(curso.sala)
        ws['A6'].font = subtitle12
        ws['A6'] = "Turno: {}".format(curso.turno_str())

        ws['A8'].border = thin_border
        ws['A8'] = "N° de Orden"
        ws['A8'].alignment = Alignment(horizontal='center',vertical='center')
        ws['B8'].border = thin_border
        ws['B8'] = "Nombre"
        ws['B8'].alignment = Alignment(horizontal='center',vertical='center')
        ws['C8'].border = thin_border
        ws['C8'] = "D.N.I."
        ws['C8'].alignment = Alignment(horizontal='center',vertical='center')
        count = 0
        number = 8
        for matricula in curso.getMatriculasAlumnos():
            count+=1
            orden = ws['A{}'.format(number+count)]
            name = ws['B{}'.format(number+count)]
            dni = ws['C{}'.format(number+count)]
            orden.border = thin_border
            orden.alignment = Alignment(horizontal='center',vertical='center')
            name.border = thin_border
            name.alignment = Alignment(horizontal='left',vertical='center')
            dni.border = thin_border
            dni.alignment = Alignment(horizontal='center',vertical='center')
            ws['A{}'.format(number+count)] = "{}".format(count).zfill(2)
            ws['B{}'.format(number+count)] = "{}, {}".format(matricula.alumno.apellidos, matricula.alumno.nombres).upper()
            ws['C{}'.format(number+count)] = "{:,}".format(int(matricula.alumno.dni)).replace(',','.')



    workbook.save(response)
    return response



def generar_xls_ministerio(request, anio):
    from openpyxl import load_workbook

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;')
    response['Content-Disposition'] = 'attachment; filename=ESCUALAS_ALUMNOS_{}.xls'.format(anio)

    workbook = load_workbook(settings.FILE_XLS_ME_CUIL)
    ws = workbook['NivelInicial']
    CODIGO_PLAN = "687196449"
    NIVEL = "INICIAL"
    COD_EMPRESA="EE1110025"
    CUE="1404229"
    CUE_ANEXO="00"
    NIVEL_EDUCATIVO="INICIAL"
    ESCUELA="AMPARO DE MARIA"
    ZONA="ZONA 1"

    cursos = Curso.objects.filter(ciclo=anio)
    #workbook = Workbook()
    #workbook.remove_sheet(workbook.active)
    count=2
    text_font = Font(name='Arial', size=10, vertAlign="baseline")
    for curso in cursos:
        for matricula in curso.getMatriculasAlumnos():
            count+=1
            ws['A{}'.format(count)] = COD_EMPRESA
            ws['B{}'.format(count)] = CUE
            ws['C{}'.format(count)] = CUE_ANEXO
            ws['D{}'.format(count)] = NIVEL_EDUCATIVO
            ws['E{}'.format(count)] = ESCUELA
            ws['F{}'.format(count)] = ZONA
            ws['G{}'.format(count)] = matricula.alumno.cuil
            ws['H{}'.format(count)] = matricula.alumno.apellidos.upper()
            ws['I{}'.format(count)] = matricula.alumno.nombres.upper()
            ws['J{}'.format(count)] = matricula.alumno.sex_type().upper()
            ws['K{}'.format(count)] = matricula.alumno.fechaDeNacimiento.strftime('%m/%d/%Y')
            ws['L{}'.format(count)] = matricula.alumno.pais.upper()
            ws['M{}'.format(count)] = CODIGO_PLAN
            ws['N{}'.format(count)] = NIVEL
            ws['O{}'.format(count)] = curso.anio
            ws['P{}'.format(count)] = curso.sala
            ws['Q{}'.format(count)] = curso.turno_str().upper()
            for l in 'ABCDEFGHIJKLMNOPQ':
                ws['{}{}'.format(l,count)].font = text_font

    workbook.save(response)
    return response



